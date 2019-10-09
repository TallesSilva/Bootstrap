from datetime import (datetime, timedelta)
from constants import (MONGO_DEFAULT_DB, MONGO_HOST, MONGO_PASS, MONGO_PORT, MONGO_USER)
from faker import Faker
from find_manage import *
from interfaces import *
from insert_manage import *
from fake_profile import *
from manage_visits import *
import random
from openpyxl import load_workbook
import logging

fake = Faker('pt_BR')
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

class excel:
    def __init__(self):
        super(excel, self).__init__()
        self.ws = []

    def insert_backlog_in_db(self):
        try:
            e = excel()
            e.load_backlog()
            row, column = e.max_row_column()
            for r in range(1, row+1):
                customer = e.read_cell(r, 1)
                supplier = e.read_cell(r, 2)
                task = e.read_cell(r, 3)
                start_date = e.read_cell(r, 4)
                ''' inserir data in mongo '''
                payload = Manage.generate_none_payload_visit(customer, supplier, start_date, task)
                f = Insert_Backlog_Payload()
                f.generate(payload)
                f.insert_to_mongo()
            return True
        except:
            return False

    def load_backlog(self):
        """ carrega um backlog em self.ws """
        try:
            print('insira o caminho para o arquivo de backlog: ')
            diretorio = input('')
            wb = load_workbook(diretorio)
            self.ws = wb.active
            return self.ws
        except Exception as log:
            print('Falha ao encontrar backlog')
            return None
    
    def read_cell(self, nrow, ncolumn):
        '''ler celula e retornar data'''
        try:
            data = self.ws.cell(nrow, ncolumn).value
            return data
        except:
            
            return None

    def max_row_column(self):
        """retorna o numero maximo de linhas e colunas"""
        try:
            ws = self.ws
            rows = ws.max_row
            columns = ws.max_column
            return rows, columns
        except:
            print("não foi possivel calcular os valores maximos")
            return None, None


if __name__ == '__main__':
    e = excel()
    e.insert_backlog_in_db()

