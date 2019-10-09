from datetime import (datetime, timedelta)
from constants import (MONGO_DEFAULT_DB, MONGO_HOST, MONGO_PASS, MONGO_PORT, MONGO_USER)
from faker import Faker
from find_manage import *
from interfaces import *
from insert_manage import *
from fake_profile import *
from manage_visits import *
import random
from openpyxl import load_workbook
import logging

fake = Faker('pt_BR')
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

class excel:
    def __init__(self):
        super(excel, self).__init__()
        self.ws = []

    def generate_timetable_with_backlog(start_create: datetime, finished_create: datetime):
        """Inicializa variaveis necessarias e lista todos os customers sem visita e cria visitas."""
        start_date = start_create
        end_date = Manage.date_sum_hour(start_date, 1)
        customer_no_have_date, customer_have_date = e.customer_list()
        print(customer_no_have_date, '\n')
        print(customer_have_date)

        '''for customer in customer_backlog:
            supplier = Getter.find_date_avaible(start_date)
            while suppliers is None:
                """Percorre as datas a procura de supplier != None."""
                start_date = Manage.date_sum_hour(start_date, 1)
                end_date = Manage.date_sum_hour(start_date, 1)
                start_date = Manage.available_date(start_date)
                suppliers = Manage.find_available_suppliers(start_date)
           
            """atualiza as condições para gerar as visitas ou não"""
            """start_date = finished_create o sistema para de gerar visitas."""
            condition_finish_create_visits = start_date >= finished_create
            conditon_continue_create_visits = start_date < finished_create
            if condition_finish_create_visits:
                break
            else:
                x = len(suppliers)
                x = random.randint(0, (x-1))
                payload = Manage.generate_available_payload_visit(customer, suppliers[x], start_date, end_date)
                Manage.insert_payload(payload)
        '''
        return True

    def customer_available(self):
        try:
            customer_backlog = Getter.get_all_backlog('customer')
            customer_no_have_date = Getter.find_all_customer_have_date('backlog', None)
            customer_have_date = [x for x in customer_backlog if x not in customer_no_have_date]
            return customer_no_have_date, customer_have_date
        except:
            return 'falha'

    def insert_backlog_in_db(self):
        try:
            e = excel()
            e.load_backlog()
            row, column = e.max_row_column()
            for r in range(1, row+1):
                customer = e.read_cell(r, 1)
                supplier = e.read_cell(r, 2)
                task = e.read_cell(r, 3)
                start_date = e.read_cell(r, 4)
                ''' inserir data in mongo '''
                payload = Manage.generate_none_payload_visit(customer, supplier, start_date, task)
                f = Insert_Backlog_Payload()
                f.generate(payload)
                f.insert_to_mongo()
            return True
        except:
            return False

    def load_backlog(self):
        """ carrega um backlog em self.ws """
        try:
            print('insira o caminho para o arquivo de backlog: ')
            diretorio = input('')
            wb = load_workbook(diretorio)
            self.ws = wb.active
            return self.ws
        except Exception as log:
            print('Falha ao encontrar backlog')
            return None
    
    def find_date_avaible(start_date):
        supplier_visits = Getter.find_all_date_supplier('time_table', )

        return supplier

    def read_cell(self, nrow, ncolumn):
        '''ler celula e retornar data'''
        try:
            data = self.ws.cell(nrow, ncolumn).value
            return data
        except:
            
            return None

    def max_row_column(self):
        """retorna o numero maximo de linhas e colunas"""
        try:
            ws = self.ws
            rows = ws.max_row
            columns = ws.max_column
            return rows, columns
        except:
            print("não foi possivel calcular os valores maximos")
            return None, None


if __name__ == '__main__':
    e = excel()
    e.insert_backlog_in_db()
    print(e.customer_available())
